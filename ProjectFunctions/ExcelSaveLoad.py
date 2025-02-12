import openpyxl
import random

from ProjectClasses.EmployeeC import Employee
from ProjectFunctions.EmployeeF import departments


# Функция для сохранения сотрудников в Excel файл
def save_employees_to_excel(filename="employees.xlsx"):
    """Сохраняет информацию о сотрудниках из словаря departments в Excel файл."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Заголовки столбцов
    headers = ["Name", "Department", "Age", "Experience", "Attentiveness",
               "Technical Literacy", "Stress Resistance", "Instruction Following",
               "Learnability", "Social Engineering Awareness", "Reporting Culture",
               "Authority Respect", "Workload", "Risk Aversion"]
    sheet.append(headers)

    for department in departments.values():
        for employee in department.employees:  # Предполагается, что у department есть атриpipбут employees, содержащий список сотрудников
            row = [employee.name, employee.department, employee.age, employee.experience,
                   employee.attentiveness, employee.technical_literacy, employee.stress_resistance,
                   employee.instruction_following, employee.learnability, employee.social_engineering_awareness,
                   employee.reporting_culture, employee.authority_respect, employee.workload,
                   employee.risk_aversion]
            sheet.append(row)

    workbook.save(filename)
    print(f"Сотрудники сохранены в файл {filename}")


# Функция для загрузки сотрудников из Excel файла
def load_employees_from_excel(filename="employees.xlsx"):
    """Загружает информацию о сотрудниках из Excel файла и создает объекты Employee.
       Возвращает список объектов Employee.
    """
    employees = []
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]] # Читаем заголовки (предполагается, что они в первой строке)

        # Проверяем, что заголовки соответствуют ожидаемым
        expected_headers = ["Name", "Department", "Age", "Experience", "Attentiveness",
                           "Technical Literacy", "Stress Resistance", "Instruction Following",
                           "Learnability", "Social Engineering Awareness", "Reporting Culture",
                           "Authority Respect", "Workload", "Risk Aversion"]
        if headers != expected_headers:
            raise ValueError("Структура Excel файла не соответствует ожидаемой.")


        for row in sheet.iter_rows(min_row=2, values_only=True):  # Начинаем со второй строки (пропускаем заголовки)
            if all(cell is None for cell in row): # Пропускаем пустые строки
                continue
            try:
                name, department, age, experience, attentiveness, technical_literacy, stress_resistance, instruction_following, learnability, social_engineering_awareness, reporting_culture, authority_respect, workload, risk_aversion = row

                # Преобразуем типы данных, если это необходимо.  Например, age и experience в int:
                age = int(age)
                experience = int(experience)

                employee = Employee(
                    name=name,
                    department=department,
                    age=age,
                    experience=experience,
                    attentiveness=attentiveness,
                    technical_literacy=technical_literacy,
                    stress_resistance=stress_resistance,
                    instruction_following=instruction_following,
                    learnability=learnability,
                    social_engineering_awareness=social_engineering_awareness,
                    reporting_culture=reporting_culture,
                    authority_respect=authority_respect,
                    workload=workload,
                    risk_aversion=risk_aversion
                )
                employees.append(employee)
            except Exception as e:
                print(f"Ошибка при обработке строки: {row}. Ошибка: {e}")


    except FileNotFoundError:
        print(f"Файл {filename} не найден.")
        return []
    except Exception as e:
        print(f"Произошла ошибка при загрузке данных из файла {filename}: {e}")
        return []

    print(f"Загружено {len(employees)} сотрудников из файла {filename}")
    return employees